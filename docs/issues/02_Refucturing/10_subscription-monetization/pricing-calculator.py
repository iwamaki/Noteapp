#!/usr/bin/env python3
"""
価格戦略計算用Excelファイル生成スクリプト
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 新しいワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "価格戦略"

# スタイル定義
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
sub_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
sub_header_font = Font(bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 列幅設定
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 15

# ====== API基準コスト ======
row = 1
ws[f'A{row}'] = "API基準コスト"
ws[f'A{row}'].font = Font(bold=True, size=14)

row += 1
ws[f'A{row}'] = "モデル"
ws[f'B{row}'] = "入力単価"
ws[f'C{row}'] = "出力単価"
ws[f'D{row}'] = "入力:出力"
ws[f'E{row}'] = "混合コスト"
for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}{row}'].fill = sub_header_fill
    ws[f'{col}{row}'].font = sub_header_font
    ws[f'{col}{row}'].border = border

row += 1
ws[f'A{row}'] = "Flash"
ws[f'B{row}'] = 0.30  # USD per 1M
ws[f'C{row}'] = 2.50
ws[f'D{row}'] = "1:3"
# 混合コスト = (入力 * 1 + 出力 * 3) / 4 * 為替レート
ws[f'E{row}'] = "=(B3*1+C3*3)/4*150"
ws[f'E{row}'].number_format = '¥#,##0'

row += 1
ws[f'A{row}'] = "Pro"
ws[f'B{row}'] = 1.25
ws[f'C{row}'] = 10.00
ws[f'D{row}'] = "1:3"
ws[f'E{row}'] = "=(B4*1+C4*3)/4*150"
ws[f'E{row}'].number_format = '¥#,##0'

row += 1
ws[f'A{row}'] = "Pro換算係数"
ws[f'E{row}'] = "=E4/E3"
ws[f'E{row}'].number_format = '0.00'
ws[f'A{row}'].font = Font(bold=True)

# ====== 単発購入（Flash） ======
row += 2
ws[f'A{row}'] = "単発購入（Flash のみ）"
ws[f'A{row}'].font = Font(bold=True, size=14)

row += 1
headers = ["商品", "価格", "トークン(M)", "APIコスト", "粗利", "利益率", "単価"]
for i, header in enumerate(headers):
    cell = ws.cell(row=row, column=i+1)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

# 単発購入データ
purchases = [
    ("初回", 300, 0.95),
    ("スモール", 300, 0.85),
    ("レギュラー", 500, 1.5),
    ("ラージ", 1000, 3.15),
]

start_row = row + 1
for i, (name, price, tokens) in enumerate(purchases):
    row += 1
    ws[f'A{row}'] = name
    ws[f'B{row}'] = price
    ws[f'B{row}'].number_format = '¥#,##0'
    ws[f'C{row}'] = tokens
    ws[f'C{row}'].number_format = '0.0'
    # APIコスト = トークン * Flash混合コスト
    ws[f'D{row}'] = f"=C{row}*$E$3"
    ws[f'D{row}'].number_format = '¥#,##0'
    # 粗利 = 価格 - APIコスト
    ws[f'E{row}'] = f"=B{row}-D{row}"
    ws[f'E{row}'].number_format = '¥#,##0'
    # 利益率 = 粗利 / 価格
    ws[f'F{row}'] = f"=E{row}/B{row}"
    ws[f'F{row}'].number_format = '0%'
    # 単価 = 価格 / トークン
    ws[f'G{row}'] = f"=B{row}/C{row}"
    ws[f'G{row}'].number_format = '¥#,##0'

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = border

# ====== サブスク（Flash + Pro） ======
row += 2
ws[f'A{row}'] = "サブスクリプション（Flash + Pro）"
ws[f'A{row}'].font = Font(bold=True, size=14)

row += 1
headers = ["プラン", "価格", "Flash(M)", "Pro(M)", "APIコスト", "粗利", "利益率", "Flash換算単価"]
for i, header in enumerate(headers):
    cell = ws.cell(row=row, column=i+1)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

# サブスクデータ
subscriptions = [
    ("Standard", 500, 1.4, 0.05),
    ("Pro", 1500, 4.5, 0.15),
    ("Premium", 3000, 9.5, 0.15),
]

for i, (name, price, flash, pro) in enumerate(subscriptions):
    row += 1
    ws[f'A{row}'] = name
    ws[f'B{row}'] = price
    ws[f'B{row}'].number_format = '¥#,##0'
    ws[f'C{row}'] = flash
    ws[f'C{row}'].number_format = '0.0'
    ws[f'D{row}'] = pro
    ws[f'D{row}'].number_format = '0.00'
    # APIコスト = Flash * Flash混合コスト + Pro * Pro混合コスト
    ws[f'E{row}'] = f"=C{row}*$E$3+D{row}*$E$4"
    ws[f'E{row}'].number_format = '¥#,##0'
    # 粗利 = 価格 - APIコスト
    ws[f'F{row}'] = f"=B{row}-E{row}"
    ws[f'F{row}'].number_format = '¥#,##0'
    # 利益率 = 粗利 / 価格
    ws[f'G{row}'] = f"=F{row}/B{row}"
    ws[f'G{row}'].number_format = '0%'
    # Flash換算単価 = 価格 / (Flash + Pro * Pro換算係数)
    ws[f'H{row}'] = f"=B{row}/(C{row}+D{row}*$E$5)"
    ws[f'H{row}'].number_format = '¥#,##0'

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{row}'].border = border

# ====== 比較表 ======
row += 2
ws[f'A{row}'] = "単価比較（高→低）"
ws[f'A{row}'].font = Font(bold=True, size=14)

row += 1
ws[f'A{row}'] = "商品/プラン"
ws[f'B{row}'] = "単価"
for col in ['A', 'B']:
    ws[f'{col}{row}'].fill = sub_header_fill
    ws[f'{col}{row}'].font = sub_header_font
    ws[f'{col}{row}'].border = border

comparison_rows = [
    ("スモール (単発)", f"=G{start_row+1}"),
    ("レギュラー (単発)", f"=G{start_row+2}"),
    ("ラージ (単発)", f"=G{start_row+3}"),
    ("Standard (サブスク)", f"=H{start_row+6}"),
    ("初回 (単発)", f"=G{start_row}"),
    ("Pro (サブスク)", f"=H{start_row+7}"),
    ("Premium (サブスク)", f"=H{start_row+8}"),
]

for name, formula in comparison_rows:
    row += 1
    ws[f'A{row}'] = name
    ws[f'B{row}'] = formula
    ws[f'B{row}'].number_format = '¥#,##0'
    for col in ['A', 'B']:
        ws[f'{col}{row}'].border = border

# ファイル保存
output_path = "/home/iwash/02_Repository/Noteapp/app/docs/pricing-calculator.xlsx"
wb.save(output_path)
print(f"✓ Excelファイルを作成しました: {output_path}")
